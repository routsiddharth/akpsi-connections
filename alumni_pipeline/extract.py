"""Read the Alumni Master Key xlsx into a flat list of person dicts.

Column positions vary per tab (2022 has an extra INDUSTRY column), so we map
*by header name*, never by fixed index.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

from . import config


@dataclass
class Person:
    first_name: str
    last_name: str
    full_name: str
    grad_year: int | None
    school: str | None = None
    industry: str | None = None
    location: str | None = None
    email_personal: str | None = None
    email_school: str | None = None
    phone: str | None = None
    linkedin: str | None = None
    post_grad_raw: str | None = None   # the messy firm string
    position_raw: str | None = None    # title / role


# Header label (normalized) -> Person field
HEADER_MAP = {
    "FIRST": "first_name",
    "LAST": "last_name",
    "YEAR": "grad_year",
    "SCHOOL": "school",
    "INDUSTRY": "industry",
    "LOCATION": "location",
    "EMAIL (PERSONAL)": "email_personal",
    "EMAIL (SCHOOL)": "email_school",
    "NUMBER": "phone",
    "LINKEDIN": "linkedin",
    "POST GRAD": "post_grad_raw",
    "POSITION": "position_raw",
}


def _norm_header(h) -> str:
    return re.sub(r"\s+", " ", str(h).strip().rstrip(":").upper())


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "N/A", "n/a", "Not specified", "TBD"):
        return None
    return s


def _phone(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—"):
        return None
    # openpyxl reads the NUMBER column as a float (4159400442.0)
    if isinstance(v, float):
        return str(int(v))
    return re.sub(r"\.0$", "", s)


def extract_people(path=None) -> list[Person]:
    path = path or config.ALUMNI_XLSX
    wb = openpyxl.load_workbook(path, data_only=True)
    people: list[Person] = []

    for tab in wb.sheetnames:
        if not re.fullmatch(r"20\d\d", tab):
            continue  # skip Legend etc.
        ws = wb[tab]

        # Build header -> column index for THIS tab. On duplicate headers (2022 has
        # two INDUSTRY cols) keep the first; the second is a stray empty column.
        col_of: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h is None:
                continue
            key = _norm_header(h)
            if key in HEADER_MAP and key not in col_of:
                col_of[key] = c

        for r in range(2, ws.max_row + 1):
            last = _clean(ws.cell(r, col_of["LAST"]).value) if "LAST" in col_of else None
            first = _clean(ws.cell(r, col_of["FIRST"]).value) if "FIRST" in col_of else None
            if not last and not first:
                continue  # blank row

            vals = {}
            for header, field_name in HEADER_MAP.items():
                if header not in col_of:
                    continue
                raw = ws.cell(r, col_of[header]).value
                vals[field_name] = _phone(raw) if field_name == "phone" else _clean(raw)

            first = vals.get("first_name") or ""
            last = vals.get("last_name") or ""
            full = re.sub(r"\s+", " ", f"{first} {last}").strip()

            gy = vals.get("grad_year")
            try:
                gy = int(float(gy)) if gy is not None else int(tab)
            except (TypeError, ValueError):
                gy = int(tab)

            people.append(
                Person(
                    first_name=first,
                    last_name=last,
                    full_name=full,
                    grad_year=gy,
                    school=vals.get("school"),
                    industry=vals.get("industry"),
                    location=vals.get("location"),
                    email_personal=vals.get("email_personal"),
                    email_school=vals.get("email_school"),
                    phone=vals.get("phone"),
                    linkedin=vals.get("linkedin"),
                    post_grad_raw=vals.get("post_grad_raw"),
                    position_raw=vals.get("position_raw"),
                )
            )

    wb.close()
    return people


if __name__ == "__main__":
    ppl = extract_people()
    print(f"Extracted {len(ppl)} people")
    for p in ppl[:5]:
        print(f"  {p.grad_year} {p.full_name:25} | {p.post_grad_raw} | {p.position_raw}")
