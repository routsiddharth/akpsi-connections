"""Read the Online Brother Profile xlsx into structured BrotherRows.

Sheet shape: 1 worksheet, ~88 brothers, divided into 7 Greek-letter pledge classes
via section header rows (col 1 = 'PSI CLASS' / 'OMEGA CLASS' / 'ALPHA ALPHA CLASS'
/ ..., other cols empty). We track the current class as we walk rows and stamp it
on each brother.

Work experience comes as free-form 'Role @ Firm' segments comma-separated.
Observed variants we handle:
    'Investment Intern @ Holocene Advisors'      -> 1 stint, role + firm
    'SWE @SpaceX @Shure'                          -> 1 role + 2 firms (2 stints)
    'Citadel'                                     -> firm only, no role
    'Incoming SA @ Goldman Sachs Asset Mgmt, ...' -> comma-split
Career Path uses mixed comma/slash separators ('Finance, Consulting' /
'Consulting/Tech' / 'Finance / pre - law').

Every normalization decision lands in data/brother_parse_log.json so weird
patterns are auditable instead of silently dropped.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field

import openpyxl

from . import config


@dataclass
class Stint:
    firm_raw: str
    group_raw: str | None = None       # rare in this format; left for completeness
    role: str | None = None


@dataclass
class BrotherRow:
    full_name: str
    first_name: str
    last_name: str
    email: str | None
    school: str | None
    grad_year: int | None
    pledge_class: str | None           # 'PSI' / 'OMEGA' / 'ALPHA ALPHA' / ...
    industries: list[str]
    work_experience: list[Stint]
    fun_fact: str | None
    source_row: int


# 'PSI CLASS' / 'ALPHA ALPHA CLASS' -> 'PSI' / 'ALPHA ALPHA'
_CLASS_SUFFIX = re.compile(r"\s+CLASS\s*$", re.IGNORECASE)
_PARENS       = re.compile(r"\([^)]*\)")
# Industries: split on comma OR slash, with optional surrounding whitespace.
_IND_SPLIT    = re.compile(r"\s*[,/]\s*")


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in {"-", "—", "n/a", "na"}:
        return None
    return s


def _is_section_header(head: list) -> bool:
    """Col 1 set, cols 2-5 all empty -> section header divider."""
    return head[0] is not None and all(v is None for v in head[1:5])


def _parse_industries(raw: str | None, log: list, row: int) -> list[str]:
    if not raw:
        return []
    parts = [p.strip() for p in _IND_SPLIT.split(raw) if p.strip()]
    # Defensive: if a stray header row leaked into the data, drop it.
    parts = [p for p in parts if p.lower() != "career path"]
    if len(parts) > 1:
        log.append({"row": row, "kind": "industries_split", "raw": raw, "into": parts})
    return parts


def _parse_work_experience(raw: str | None, log: list, row: int) -> list[Stint]:
    """Split each comma-separated segment, then split each segment on '@'.
    Strip parentheticals (often dates / 'Summer 2024')."""
    if not raw:
        return []
    raw = str(raw).strip()
    if not raw or raw.lower() == "internship/work/research experience":
        return []
    stints: list[Stint] = []
    for seg in raw.split(","):
        seg = seg.strip()
        if not seg:
            continue
        if "@" in seg:
            role_part, *firm_parts = [p.strip() for p in seg.split("@")]
            role = role_part or None
            for f in firm_parts:
                f_clean = _PARENS.sub("", f).strip(" -")
                if f_clean:
                    stints.append(Stint(firm_raw=f_clean, role=role))
        else:
            # No '@' separator — treat the whole segment as a firm name, log it.
            f_clean = _PARENS.sub("", seg).strip(" -")
            if f_clean:
                stints.append(Stint(firm_raw=f_clean, role=None))
                log.append({"row": row, "kind": "no_at_separator",
                            "raw": seg, "treated_as": "firm_only"})
    return stints


def parse(path=None) -> list[BrotherRow]:
    path = path or config.BROTHER_XLSX
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    # Header row 1 -> column-index map, located by label name (not fixed position).
    HEADER_MAP = {
        "first name": "first_name",
        "last name": "last_name",
        "e-mail": "email",
        "school": "school",
        "grad year": "grad_year",
        "career path": "career",
        "internship/work/research experience": "work_exp",
        "interesting/fun fact": "fun_fact",
    }
    col_of: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        key = str(h).strip().lower()
        if key in HEADER_MAP:
            col_of[HEADER_MAP[key]] = c

    log: list = []
    rows: list[BrotherRow] = []
    current_class: str | None = None

    def cv(r: int, name: str):
        return ws.cell(r, col_of[name]).value if name in col_of else None

    for r in range(2, ws.max_row + 1):
        head = [ws.cell(r, c).value for c in range(1, 6)]
        if _is_section_header(head):
            current_class = _CLASS_SUFFIX.sub("", str(head[0]).strip()).upper()
            log.append({"row": r, "kind": "section_header", "pledge_class": current_class})
            continue

        first = _clean(cv(r, "first_name"))
        last  = _clean(cv(r, "last_name"))
        if not first and not last:
            continue
        full = re.sub(r"\s+", " ", f"{first or ''} {last or ''}").strip()

        gy_raw = _clean(cv(r, "grad_year"))
        try:
            grad_year = int(float(gy_raw)) if gy_raw else None
        except (TypeError, ValueError):
            grad_year = None
            log.append({"row": r, "kind": "grad_year_parse_failed",
                        "raw": gy_raw, "name": full})

        rows.append(BrotherRow(
            full_name=full,
            first_name=first or "",
            last_name=last or "",
            email=_clean(cv(r, "email")),
            school=_clean(cv(r, "school")),
            grad_year=grad_year,
            pledge_class=current_class,
            industries=_parse_industries(_clean(cv(r, "career")), log, r),
            work_experience=_parse_work_experience(_clean(cv(r, "work_exp")), log, r),
            fun_fact=_clean(cv(r, "fun_fact")),
            source_row=r,
        ))

    wb.close()

    config.DATA_DIR.mkdir(exist_ok=True)
    config.PARSE_LOG_FILE.write_text(json.dumps(log, indent=2))

    return rows


if __name__ == "__main__":
    bs = parse()
    by_class: dict[str, int] = {}
    for b in bs:
        by_class[b.pledge_class or "?"] = by_class.get(b.pledge_class or "?", 0) + 1
    print(f"Parsed {len(bs)} brothers across {len(by_class)} pledge classes:")
    for k, v in sorted(by_class.items()):
        print(f"  {k}: {v}")
    print("\nFirst 3:")
    for b in bs[:3]:
        print(f"  {b.full_name} ({b.pledge_class}, {b.grad_year}, school={b.school})")
        print(f"    industries: {b.industries}")
        print(f"    work: {[(s.role, s.firm_raw) for s in b.work_experience]}")
