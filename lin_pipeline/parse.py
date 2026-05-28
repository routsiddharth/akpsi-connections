"""Read the Lin Trees xlsx into normalized lineage rows + roster entries.

Per-tab quirks live in config.LINEAGE_TABS (header offset, column positions);
unlisted tabs fall through to the default two-column layout. The single
'TAKING LITTLES 2026' roster tab is parsed separately because it has a wider
schema (Name / Year / Industry / Lin) — its Year column is a clean grad_year
source for new undergrads.

Cell formatting carries chapter status:
    bold     -> current brother
    italic   -> alumnus
    '*' suff -> abroad / inactive
Stripped from the name for matching but captured as metadata.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

from . import config


@dataclass
class Markers:
    bold: bool = False
    italic: bool = False
    starred: bool = False


@dataclass
class LineageRow:
    little_raw: str          # cleaned name string; also the person_aliases key
    big_raw: str
    little_markers: Markers
    big_markers: Markers
    source_tab: str


@dataclass
class RosterEntry:
    name_raw: str
    grad_year: int | None
    industry: str | None
    family: str | None
    markers: Markers = field(default_factory=Markers)


_PARENTHETICAL = re.compile(r"\([^)]*\)")  # "Daniel (Yoonsuk) Choi" -> "Daniel  Choi"
_WS = re.compile(r"\s+")


def _clean_name(v) -> tuple[str | None, Markers]:
    """Return (cleaned_name, markers). Handles None, empty, and placeholder cells."""
    if v is None:
        return None, Markers()
    s = str(v).strip()
    if not s or s in {"-", "—"}:
        return None, Markers()
    markers = Markers()
    if s.endswith("*"):
        markers.starred = True
        s = s.rstrip("* ").strip()
    s = _WS.sub(" ", _PARENTHETICAL.sub(" ", s)).strip()
    return (s or None), markers


def _fold_font(cell, markers: Markers) -> Markers:
    if cell.font:
        if cell.font.bold:
            markers.bold = True
        if cell.font.italic:
            markers.italic = True
    return markers


def _read_lineage(ws, cfg: dict, tab: str) -> list[LineageRow]:
    little_col, big_col = cfg["little_col"], cfg["big_col"]
    first_row = (cfg["header_row"] + 1) if cfg["header_row"] is not None else 1
    rows: list[LineageRow] = []
    for r in range(first_row, ws.max_row + 1):
        l_cell, b_cell = ws.cell(r, little_col), ws.cell(r, big_col)
        l_name, l_markers = _clean_name(l_cell.value)
        b_name, b_markers = _clean_name(b_cell.value)
        if not l_name or not b_name:
            continue                          # incomplete row — skip silently
        rows.append(LineageRow(
            little_raw=l_name, big_raw=b_name,
            little_markers=_fold_font(l_cell, l_markers),
            big_markers=_fold_font(b_cell, b_markers),
            source_tab=tab,
        ))
    return rows


def _read_roster(ws) -> list[RosterEntry]:
    """TAKING LITTLES 2026: header at row 1, locate cols by name (tolerates shuffles)."""
    col_of: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        key = str(h).strip().lower()
        if key in {"name", "year", "industry", "lin"}:
            col_of.setdefault(key, c)
    if "name" not in col_of:
        return []

    out: list[RosterEntry] = []
    for r in range(2, ws.max_row + 1):
        n_cell = ws.cell(r, col_of["name"])
        name, markers = _clean_name(n_cell.value)
        if not name:
            continue
        gy_raw = ws.cell(r, col_of["year"]).value if "year" in col_of else None
        try:
            gy = int(float(gy_raw)) if gy_raw is not None else None
        except (TypeError, ValueError):
            gy = None
        ind_v = ws.cell(r, col_of["industry"]).value if "industry" in col_of else None
        fam_v = ws.cell(r, col_of["lin"]).value      if "lin"      in col_of else None
        out.append(RosterEntry(
            name_raw=name, grad_year=gy,
            industry=str(ind_v).strip() if ind_v else None,
            family=str(fam_v).strip() if fam_v else None,
            markers=_fold_font(n_cell, markers),
        ))
    return out


def parse(path=None) -> tuple[list[LineageRow], list[RosterEntry]]:
    path = path or config.LINTREES_XLSX
    wb = openpyxl.load_workbook(path, data_only=True)
    lineage: list[LineageRow] = []
    roster: list[RosterEntry] = []

    for tab in wb.sheetnames:
        if tab == config.ROSTER_TAB:
            roster = _read_roster(wb[tab])
            continue
        cfg = config.LINEAGE_TABS.get(tab, config.DEFAULT_LINEAGE)
        lineage.extend(_read_lineage(wb[tab], cfg, tab))

    wb.close()
    return lineage, roster


if __name__ == "__main__":
    edges, roster = parse()
    tabs = sorted({e.source_tab for e in edges})
    print(f"{len(edges)} lineage edges across {len(tabs)} tabs: {tabs}")
    print(f"{len(roster)} roster entries (TAKING LITTLES 2026)")
    for e in edges[:5]:
        print(f"  [{e.source_tab}] little={e.little_raw} <- big={e.big_raw}")
